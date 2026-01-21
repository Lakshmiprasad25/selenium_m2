import openpyxl

def read_login_data(file_path):
    workbook=openpyxl.load_workbook(file_path)
    sheet=workbook["Sheet1"]
    rows=sheet.max_row
    data=[]
    for r in range(2,rows+1):
        username=sheet.cell(r,1).value
        password=sheet.cell(r,2).value
        data.append((username, password))
    return data
